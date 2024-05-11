import openpyxl
import os

file_path = 'N_Queen.xlsx'

# Check if the file exists
if os.path.exists(file_path):
    Workbook1 = openpyxl.load_workbook(file_path)
    worksheet = Workbook1['Sheet1']
else:
    # Create a new workbook
    Workbook1 = openpyxl.Workbook()
    worksheet = Workbook1.active
    worksheet.title = 'Sheet1'

print("\t\t\t\t\t\t\t\t\t\t\t\t\tWELLCOME TO N-QUEEN SOLVER!")
total = int(input("\t\tYour given number should be greater than 3. The best range is 4_10.\n\t\tHow many queens you want to place?\n\t\t\t\t\t\t"))
print()
while total < 4:
    print("\t\tInvalid number !")
    total = int(input("\t\tQueens should be from 4-10\n\t\t\t\tInput again:\t"))
    print()
t = total
mylist = []
while t > 0:
    mylist.append("Queen")
    t -= 1
counter = 0
colum = 0


class N_Queen:
    def column_check(self,r, c):
        flag = 0
        a = total - 1
        while a > 0:
            r = r % total
            r += 1
            if worksheet.cell(row=r, column=c).value != None:
                flag = 1
            a -= 1
        return flag

    def column_row_check(self,r, c):
        flag = 0
        while r > 0 and c > 0 and c <= total and r <= total:
            if worksheet.cell(row=r, column=c).value != None:
                flag = 1
            r -= 1
            c -= 1
        return flag

    def column_row_check1(self,r, c):
        flag = 0
        while c > 0 and r > 0 and c <= total and r <= total:
            if worksheet.cell(row=r, column=c).value != None:
                flag = 1
            r -= 1
            c += 1
        return flag

    def backtraversal(self,r):
        for j in range(0, total):
            if worksheet.cell(row=r, column=j + 1).value != None:
                worksheet.cell(row=r, column=j + 1).value = None
                return j


for i in mylist:
    if counter<=0 :
        worksheet.cell(row=counter+1, column=colum+1).value = i
    else:
        worksheet.cell(row=counter+1,column=colum+1).value = i
        a=N_Queen().column_check(counter+1,colum+1)
        b=N_Queen().column_row_check(counter,colum)
        c=N_Queen().column_row_check1(counter,colum+2)
        if a==1 or b==1 or c==1:
            worksheet.cell(row=counter + 1, column=colum + 1).value = None
            if colum>total-2:
                y=N_Queen().backtraversal(counter)
                colum=y+1
                counter-=1
                mylist.append("Queen")
                u=counter
                while colum >= total:
                    w=N_Queen().backtraversal(u)
                    colum = w + 1
                    counter -= 1
                    mylist.append("Queen")
            else:
                colum+=1
            counter-=1
            mylist.append("Queen")
        else:
            colum=0
    counter += 1
listy=[]
for k in range(1,total+1):
    for l in range(1,total+1):
        listy.append(worksheet.cell(row=k,column=l).value)
m=1
for p in listy:
    if p==None:
        print(0,end="      ")
    else:
        print(p, end="  ")
    if m==total:
        print("\n")
        m-=total
    m+=1
for k in range(1,total+1):
    for l in range(1,total+1):
        worksheet.cell(row=k,column=l).value = None
